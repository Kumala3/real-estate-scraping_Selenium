import time
from tqdm import tqdm
import openpyxl
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException

regionId = 2  # This ID belongs only to the Tomsk region!
pages = int(input("Enter how many realtor pages you want to parse: "))


def get_url(regId, page):
    """
    The  get_url  function takes two arguments:  regId  (region ID) and  page  (page number).
    It is used to generate the URL address for a page on the website www.cian.ru, where information about realtors is displayed.
    """
    return f"https://www.cian.ru/realtors/?regionId={regId}&page={page}"


def get_urls_rielrts(driver, regionIp, pages):
    """
     The  get_urls_rielrts  function is designed to retrieve a list of URLs from a Selenium driver.
        It takes the driver, region IP address, and number of pages as arguments.
        Inside the function, there is a loop that iterates over the specified number of pages.
        Each page's URL is generated using the  get_url  function and then the driver navigates to that URL.
    """
    links = []

    try:
        for index in range(pages):
            url = f"https://www.cian.ru/realtors/?regionId={regionIp}&page={index+1}"
            driver.get(url)
        for elem in driver.find_elements(By.CLASS_NAME, "_9400a595a7--container--J25nK"):
            try:
                driver.execute_script("arguments[0].click();", elem)
                driver.switch_to.window(driver.window_handles[1])

                links.append(driver.current_url)
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
            except TimeoutException:
                time.sleep(10)
                driver.refresh()
                driver.execute_script("arguments[0].click();", elem)
                driver.switch_to.window(driver.window_handles[1])

                links.append(driver.current_url)
                driver.switch_to.window(driver.window_handles[0])
            except:
                time.sleep(10)
                driver.refresh()
                driver.execute_script("arguments[0].click();", elem)
                driver.switch_to.window(driver.window_handles[1])

                links.append(driver.current_url)
                driver.switch_to.window(driver.window_handles[0])

            WebDriverWait(driver, 8).until(ec.url_to_be(driver.current_url))


    except Exception as ex:
        print(type(ex).__name__)

    return links


def get_info_personal(driver, links):
    progress_bar = tqdm(total=len(links), desc="Парсинг данных пользователей", unit="пользователей")

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = "Ссылка на профиль"
    sheet.cell(row=1, column=2).value = "Имя"
    sheet.cell(row=1, column=3).value = "Электронная почта"
    sheet.cell(row=1, column=4).value = "Номер телефона"
    sheet.cell(row=1, column=5).value = "Опыт работы"
    sheet.cell(row=1, column=6).value = "Опыт работы на Циан"
    sheet.cell(row=1, column=7).value = "В работе"
    sheet.cell(row=1, column=8).value = "Звезды"
    sheet.cell(row=1, column=9).value = "Отзывы"
    sheet.cell(row=1, column=10).value = "'Документы"
    sheet.cell(row=1, column=11).value = "О себе"
    sheet.cell(row=1, column=12).value = "Специализация"
    sheet.cell(row=1, column=13).value = "Регион работы"
    sheet.cell(row=1, column=14).value = "Агентство"

    for link in links:
        row = sheet.max_row + 1
        driver.get(link)
        time.sleep(0.5)
        name = driver.find_element(By.CLASS_NAME, "_3ea6fa5da8--name--JPPsh").text

        element = driver.find_elements(By.CSS_SELECTOR, "div._3ea6fa5da8--socnetwork--Q6ec4")

        phone_number = ''
        mail = ''

        if len(element) == 3:
            try:
                phone_number_element = driver.find_element(By.CSS_SELECTOR, 'div._3ea6fa5da8--phones_minimized--XieZH')
                driver.execute_script("arguments[0].click();", phone_number_element)

                WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="realtor-contacts"]/div/div[4]/div/ul/li/a')))
                phone_number = driver.find_element(By.XPATH, '//*[@id="realtor-contacts"]/div/div[4]/div/ul/li/a').text
                mail = element[1].text
            except NoSuchElementException:
                phone_number += 'Отсутствует'
                mail += 'Отсутствует'

        if len(element) == 2:
            try:
                phone_number_element = driver.find_element(By.CSS_SELECTOR, 'div._3ea6fa5da8--phones_minimized--XieZH')
                driver.execute_script("arguments[0].click();", phone_number_element)

                WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="realtor-contacts"]/div/div[3]/div/ul/li/a')))
                phone_number = driver.find_element(By.XPATH, '//*[@id="realtor-contacts"]/div/div[3]/div/ul/li/a').text
                mail = element[1].text

            except NoSuchElementException:
                phone_number += 'Отсутствует'
                mail += 'Отсутствует'

        if len(element) == 1:
            try:
                phone_number_element = WebDriverWait(driver, 10).until(
                    ec.element_to_be_clickable((By.CSS_SELECTOR, 'div._3ea6fa5da8--phones_minimized--XieZH'))
                )

                driver.execute_script("arguments[0].click();", phone_number_element)

                WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="realtor-contacts"]/div/div[2]/div/ul/li/a')))
                phone_number = driver.find_element(By.XPATH, '//*[@id="realtor-contacts"]/div/div[2]/div/ul/li/a').text
                mail = element[0].text

            except NoSuchElementException:
                phone_number += 'Отсутствует'
                mail += 'Отсутствует'

        if len(element) == 0:
            try:
                phone_number_element = driver.find_element(By.CSS_SELECTOR, 'div._3ea6fa5da8--phones_minimized--XieZH')

                driver.execute_script("arguments[0].click();", phone_number_element)

                WebDriverWait(driver, 10).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="realtor-contacts"]/div/div/div/ul/li/a')))
                phone_number = driver.find_element(By.XPATH, '//*[@id="realtor-contacts"]/div/div/div/ul/li/a').text
                mail += 'Отсутствует'
            except NoSuchElementException:
                phone_number += 'Отсутствует'
                mail += 'Отсутствует'

        elements = driver.find_elements(By.CSS_SELECTOR, "div._3ea6fa5da8--counters-item--kwGtk")

        experience = ''

        if elements[0].find_element(By.CSS_SELECTOR, "span._3ea6fa5da8--color_black_100--kPHhJ._3ea6fa5da8--lineHeight_22px--bnKK9._3ea6fa5da8--fontWeight_bold--ePDnv._3ea6fa5da8--fontSize_16px--RB9YW._3ea6fa5da8--display_block--pDAEx._3ea6fa5da8--text--g9xAG._3ea6fa5da8--text_letterSpacing__normal--xbqP6").text == 'не указан':
            experience += 'Опыт работы не указан'
        else:
            experience = elements[0].find_element(By.CSS_SELECTOR, "span._3ea6fa5da8--color_black_100--kPHhJ._3ea6fa5da8--lineHeight_22px--bnKK9._3ea6fa5da8--fontWeight_bold--ePDnv._3ea6fa5da8--fontSize_16px--RB9YW._3ea6fa5da8--display_block--pDAEx._3ea6fa5da8--text--g9xAG._3ea6fa5da8--text_letterSpacing__normal--xbqP6").text

        on_cian = elements[1].find_element(By.CSS_SELECTOR, "span._3ea6fa5da8--color_black_100--kPHhJ._3ea6fa5da8--lineHeight_22px--bnKK9._3ea6fa5da8--fontWeight_bold--ePDnv._3ea6fa5da8--fontSize_16px--RB9YW._3ea6fa5da8--display_block--pDAEx._3ea6fa5da8--text--g9xAG._3ea6fa5da8--text_letterSpacing__normal--xbqP6").text

        objects = ''

        if elements[2].find_element(By.CSS_SELECTOR, "div._3ea6fa5da8--counters-item--kwGtk > span").text == 'нет объектов':
            objects += '0'
        else:
            objects = elements[2].find_element(By.CSS_SELECTOR, "div._3ea6fa5da8--counters-item--kwGtk > span").text

        dokument = driver.find_elements(By.CSS_SELECTOR, 'div[data-name="ApprovalNew"]')

        dokument_verif = ''
        if len(dokument) == 0:
            dokument_verif += "Документы агента не проверены"
        else:
            dokument_verif += "Документы агента проверены"

        rating_rewiews = ''

        if driver.find_element(By.CSS_SELECTOR, "div._3ea6fa5da8--rating-desctiption--HgRir").text == 'Нет оценок и отзывов':
            rating_rewiews += '0,0'
        else:
            rating_rewiews = driver.find_element(By.CSS_SELECTOR, "div._3ea6fa5da8--rating-desctiption--HgRir").text

        if rating_rewiews == '0,0':
            rating = rating_rewiews.split(',')[0]
            rewiews = rating_rewiews.split(',')[1]
        else:
            rating = rating_rewiews.split('・')[0]
            rewiews = rating_rewiews.split('・')[1].replace('отзыв', '')

        about_me = ''
        try:
            gg = driver.find_element(By.XPATH, '//*[@id="realtor-reviews-frontend"]/div/div[2]/main/section[1]/div[3]/div[1]').text
            skun = gg.split()

            if skun[0] == 'О':
                cebe = ' '.join(skun[0:2]) + ':\n'
                information_me = ' '.join(skun[2:]).lstrip()
                about_me = ' '.join([cebe, information_me]).lstrip()

            if skun[0] == 'Специализация':
                spec = ' '.join(skun[0:1]) + ':\n'
                information_spec = ' '.join(skun[1:]).lstrip()
                about_me = ' '.join([spec, information_spec]).lstrip()

            if skun[0] == 'Регион':
                reg_wok = ' '.join(skun[0:2]) + ':\n'
                information_reg_work = ' '.join(skun[2:]).lstrip()
                about_me = ' '.join([reg_wok, information_reg_work]).lstrip()

            if skun[0] == 'Агентство':
                ag = ' '.join(skun[0:1]) + ':\n'
                information_agency = ' '.join(skun[1:]).lstrip()
                about_me = ' '.join([ag, information_agency]).lstrip()


        except NoSuchElementException:
            about_me += "Отсутствует"

        specialization = ''
        try:
            go = driver.find_element(By.XPATH, '//*[@id="realtor-reviews-frontend"]/div/div[2]/main/section[1]/div[3]/div[2]').text

            skj = go.split()

            if skj[0] == 'Специализация':
                spec = ' '.join(skj[0:1]) + ':\n'
                information_spec = ' '.join(skj[1:]).lstrip()
                specialization = ' '.join([spec, information_spec]).lstrip()

            if skj[0] == 'Регион':
                reg_wok = ' '.join(skj[0:2]) + ':\n'
                information_reg_work = ' '.join(skj[2:]).lstrip()
                specialization = ' '.join([reg_wok, information_reg_work])

            if skj[0] == 'Агентство':
                ag = ' '.join(skj[0:1]) + ':\n'
                information_agency = ' '.join(skj[1:]).lstrip()
                specialization = ' '.join([ag, information_agency])


        except NoSuchElementException:
            specialization += 'Отсутствует'

        region_work = ''
        try:
            ab = driver.find_element(By.XPATH, '//*[@id="realtor-reviews-frontend"]/div/div[2]/main/section[1]/div[3]/div[3]').text
            fokus = ab.split()

            if fokus[0] == 'Регион':
                reg_wok = ' '.join(fokus[0:2]) + ':\n'
                information_reg_work = ' '.join(fokus[2:])
                region_work = ' '.join([reg_wok, information_reg_work]).lstrip()

            if fokus[0] == 'Агентство':
                ag = ' '.join(fokus[0:1]) + ':\n'
                information_agency = ' '.join(fokus[1:])
                region_work = ' '.join([ag, information_agency]).lstrip()

        except NoSuchElementException:
            region_work += 'Отсутствует'

        agency = ''
        try:
            agency = driver.find_element(By.XPATH, '//*[@id="realtor-reviews-frontend"]/div/div[2]/main/section[1]/div[3]/div[4]').text

        except NoSuchElementException:
            agency += 'Отсутсвует или уже присутсвует в верхних елементах'

        sheet.cell(row=row, column=1).value = link
        sheet.cell(row=row, column=2).value = name
        sheet.cell(row=row, column=3).value = mail
        sheet.cell(row=row, column=4).value = phone_number
        sheet.cell(row=row, column=5).value = experience
        sheet.cell(row=row, column=6).value = on_cian
        sheet.cell(row=row, column=7).value = objects
        sheet.cell(row=row, column=8).value = rating
        sheet.cell(row=row, column=9).value = rewiews
        sheet.cell(row=row, column=10).value = dokument_verif
        sheet.cell(row=row, column=11).value = about_me
        sheet.cell(row=row, column=12).value = specialization
        sheet.cell(row=row, column=13).value = region_work
        sheet.cell(row=row, column=14).value = agency

        workbook.save('agent_data.xlsx')
        progress_bar.update(1)
    progress_bar.close()


def main():
    options = ChromeOptions()
    # options.add_argument("--headless")
    options.add_argument("--start_maximized")

    driver = webdriver.Chrome(options=options)

    try:
        links = get_urls_rielrts(driver, regionId, pages)
        data = get_info_personal(driver, links)

    except Exception as ex:
        print(ex)
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
